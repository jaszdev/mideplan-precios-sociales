from openpyxl import load_workbook
from enum import Enum

PRODUCTS_PATH = '.\\files\\input_files\\products_updated.xlsx'

RAZON_PRECIO_SOCIAL_DIVISA = 1.0645
EXCHANGE_RATE = 504 # PENDIENTE: CALCULAR TIPO_DE_CAMBIO

class PRODUCTS_COLUMNS(Enum):
    """
    Columnas en la hoja de productos
    """
    # Informacion de Producto
    NAME = 'A'
    CODE = 'B'
    TYPE = 'C'
    
    # CIF y FOB
    CF_MNT_IMP = 'D'
    CF_MNT_SIMP = 'E'
    CF_MNT_EXP = 'F'
    CF_MNT_DEXP = 'G'

    CF_MT_IMP = 'H'
    CF_MT_SIMP = 'I'
    CF_MT_EXP = 'J'
    CF_MT_DEXP = 'K'
    
    # VALORES DE MERCADO
    VM_MT_IMP = 'L'
    VM_MT_EXP = 'M'
    VM_MT_SIMP = 'N'
    VM_MT_DESX = 'O'

    VM_MNT_IMP = 'P'
    VM_MNT_EXP = 'Q'
    VM_MNT_SIMP = 'R'
    VM_MNT_DESX = 'S'

    VM_BNT = 'T'

class RESULTS_COLUMNS(Enum):
    """
    Columnas de la hoja de resultados.
    """
    # INPUT - Informacion de Producto
    NAME = 'A'
    PRICE = 'B'
    CURRENCY = 'C'
    CODE = 'D'
    MARGIN_TYPE = 'E'
    DISTRIBUTION_MARGIN = 'F'

    # OUTPUT - Calculo del precio social
    FE =  'G' # Factor Especifico
    PS = 'H' # Precio Social

class CURRENCY(Enum):
    COLONES = 'Colones'
    DOLARES = 'Dólares'

class TYPE(Enum):
    T = 'Transable'
    NT = 'No Transable'

class MARGIN_TYPE(Enum):
    IMP = 'Importable'
    SIMP = 'Sust. Importaciones'
    EXP = 'Exportable'
    DESX = 'Desv. Exportaciones'


# PARAMETROS
METHOD = 1 # 1, 2 o 3
PHASE = 1 # 1 para Prefactibilidad, 2 para Factibilidad

INPUT_PATH = '.\\files\\inputs\\input1.xlsx'
OUTPUT_PATH = '.\\files\\output\\result1.xlsx'

# CARGAR HOJA DE BIENES T Y NT
p_ws = load_workbook(filename=PRODUCTS_PATH, data_only=True).active
input_wb = load_workbook(filename=INPUT_PATH, data_only=True)

def main():
    ws = input_wb.active

    # AÑADIR COLUMNAS DE RESULTADOS
    ws[f'{RESULTS_COLUMNS.FE.value}1'] = 'Factor Específico'
    ws[f'{RESULTS_COLUMNS.PS.value}1'] = 'Precio Social'

    row = 2
    price = ws[f'{RESULTS_COLUMNS.PRICE.value}{row}'].value
    while(price != None and price != ''):

        M1_Compute_SP(ws, row)

        row += 1
        price = ws[f'{RESULTS_COLUMNS.PRICE.value}{row}'].value


def M1_Compute_SP(ws, row):
    if PHASE == 1:
        code = ws[f'{RESULTS_COLUMNS.CODE.value}{row}'].value
        currency = ws[f'{RESULTS_COLUMNS.CURRENCY.value}{row}'].value
        marginType = ws[f'{RESULTS_COLUMNS.MARGIN_TYPE.value}{row}'].value
        
        #### PENDING: CHECK FOR ERRORS

        fe = getSpecificFactor(METHOD, PHASE, code, currency, marginType)

        # WRITE FE
        if (fe != TYPE.NT.value): # TRANSABLES
            fe_coord = f'{RESULTS_COLUMNS.FE.value}{row}'
            ws[fe_coord] = fe

            # WRITE PS
            ps_coord = f'{RESULTS_COLUMNS.PS.value}{row}'
            if fe != '-':
                rowPrice_coord = f'{RESULTS_COLUMNS.PRICE.value}{row}'
                ws[ps_coord] = f'={rowPrice_coord}*{fe_coord}'
            elif fe == '-':
                ws[ps_coord] = 'ERROR. Factor Específico no encontrado.'
        elif fe == TYPE.NT.value: # NO TRANSABLES
            fe_coord = f'{RESULTS_COLUMNS.FE.value}{row}'
            ws[fe_coord] = '-'

            # WRITE PS
            ps_coord = f'{RESULTS_COLUMNS.PS.value}{row}'
            rowPrice_coord = f'{RESULTS_COLUMNS.PRICE.value}{row}'
            ws[ps_coord] = f'={rowPrice_coord}'

    else:
        pass

def getSpecificFactor(method, phase, code, currency, marginType):
    # METODO #1 - PREFACTIBILIDAD - Disponible en excel
    if method == 1 and phase == 1 and code != None: 
        int_code = int(code[2:5]) # PARSE (NP[CODE NUMBER] - PRODUCT NAME) -> CODE NUMBER
        code_row = int_code + 1

        type_coord = f'{PRODUCTS_COLUMNS.TYPE.value}{code_row}'
        p_type = p_ws[type_coord].value

        if p_type == TYPE.NT.value:
            return TYPE.NT.value

        if (currency == CURRENCY.COLONES.value):
            if (marginType == MARGIN_TYPE.IMP.value):
                coord = f'{PRODUCTS_COLUMNS.VM_MT_IMP.value}{code_row}'
                return p_ws[coord].value
            elif (marginType == MARGIN_TYPE.SIMP.value):
                coord = f'{PRODUCTS_COLUMNS.VM_MT_SIMP.value}{code_row}'
                return p_ws[coord].value
            elif (marginType == MARGIN_TYPE.EXP.value):
                coord = f'{PRODUCTS_COLUMNS.VM_MT_EXP.value}{code_row}'
                return p_ws[coord].value
            elif (marginType == MARGIN_TYPE.DESX.value):
                coord = f'{PRODUCTS_COLUMNS.VM_MT_DESX.value}{code_row}'
                return p_ws[coord].value
        else: # DOLARES
            if (marginType == MARGIN_TYPE.IMP.value):
                coord = f'{PRODUCTS_COLUMNS.CF_MT_IMP.value}{code_row}'
                return p_ws[coord].value
            elif (marginType == MARGIN_TYPE.SIMP.value):
                coord = f'{PRODUCTS_COLUMNS.CF_MT_SIMP.value}{code_row}'
                return p_ws[coord].value
            elif (marginType == MARGIN_TYPE.EXP.value):
                coord = f'{PRODUCTS_COLUMNS.CF_MT_EXP.value}{code_row}'
                return p_ws[coord].value
            elif (marginType == MARGIN_TYPE.DESX.value):
                coord = f'{PRODUCTS_COLUMNS.CF_MT_DEXP.value}{code_row}'
                return p_ws[coord].value   
    # METODO #1 - PREFACTIBILIDAD - No disponible en excel
    elif method == 1 and phase == 1 and code == None:
        if (currency == CURRENCY.COLONES.value):
            return RAZON_PRECIO_SOCIAL_DIVISA
        else: # DOLARES
            return f'={RAZON_PRECIO_SOCIAL_DIVISA}*{EXCHANGE_RATE}'

    return -1

main()

input_wb.save(OUTPUT_PATH)